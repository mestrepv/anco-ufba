"""Recupera DOIs perdidos na revisão cruzando com a base referencial original.

Compara `base-anco-revisada.xlsx` (sem DOI estruturado) com
`base-referencial-em-analise-cognitiva.xlsx` (884 DOIs canônicos) e produz
um JSON de mapeamento `{titulo_normalizado, ano, periodico} -> doi`.

Critérios de match, por ordem decrescente de confiança:
  1. Título + ano + periódico (todos normalizados)
  2. Título + ano
  3. Título apenas (com SequenceMatcher >= 0.92)

Saída inclui o nível de confiança para que a aplicação no banco
possa diferenciar matches fortes (aplicar direto) de fracos (revisar).

Uso:
  python tools/recuperar_dois_referencial.py \\
    /app/base-anco-revisada.xlsx \\
    /app/base-referencial-em-analise-cognitiva.xlsx \\
    /app/dois_recuperados.json
"""

from __future__ import annotations

import datetime
import json
import re
import sys
import unicodedata
from collections import Counter
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl

# Regex permissiva: DOI canônico em qualquer posição da string
DOI_RE = re.compile(r"\b(10\.\d{4,9}/[-._;()/:A-Z0-9]+)", re.IGNORECASE)


def normalizar_titulo(s: str | None) -> str:
    """Lowercase, sem acentos, sem pontuação periférica, espaços colapsados."""
    if not s:
        return ""
    s = str(s)
    # Remove acentos
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.lower().strip()
    # Strip de aspas tipográficas e pontuação final
    s = re.sub(r"^[\s'\"`'‘’“”]+|[\s'\"`'‘’“”.,;:]+$", "", s)
    # Colapsa espaços
    s = re.sub(r"\s+", " ", s)
    return s


def normalizar_ano(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        a = int(float(v))
        return a if 1900 <= a <= 2030 else None
    except (TypeError, ValueError):
        return None


def normalizar_periodico(s: str | None) -> str:
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def extrair_doi(raw: str | None) -> str | None:
    """Retorna DOI canônico em lowercase, ou None."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s in ("-", "—"):
        return None
    # Strip de prefixo "DOI:"
    s = re.sub(r"^(?i:doi)\s*:\s*", "", s)
    m = DOI_RE.search(s)
    if not m:
        return None
    doi = m.group(1).rstrip(".,;)").lower()
    return doi


def carregar_revisada(path: Path) -> list[dict]:
    """Retorna [{linha_xlsx, titulo, ano, periodico, doi_atual}]."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out = []
    for i, r in enumerate(rows[2:], start=3):  # pula VOLTAR + cabeçalho
        if not r or r[0] is None:
            continue
        out.append(
            {
                "linha_xlsx": i,
                "titulo_orig": r[0] or "",
                "titulo": normalizar_titulo(r[0]),
                "ano": normalizar_ano(r[3]),
                "periodico": normalizar_periodico(r[4]),
            }
        )
    return out


def carregar_referencial(path: Path) -> list[dict]:
    """Retorna [{linha_xlsx, titulo, ano, periodico, doi}] da aba 'Base'."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Base"]
    rows = list(ws.iter_rows(values_only=True))
    # Cabeçalho na linha 1; dados a partir da linha 2
    out = []
    for i, r in enumerate(rows[1:], start=2):
        if not r or len(r) < 10:
            continue
        doi = extrair_doi(r[1])
        titulo = normalizar_titulo(r[3])
        if not titulo:
            continue
        out.append(
            {
                "linha_xlsx": i,
                "titulo_orig": r[3] or "",
                "titulo": titulo,
                "ano": normalizar_ano(r[4]),
                "periodico": normalizar_periodico(r[9]),
                "doi": doi,
            }
        )
    return out


def indexar_referencial(refs: list[dict]) -> dict[str, list[dict]]:
    """Index by título normalizado (lista — pode haver duplicatas)."""
    idx: dict[str, list[dict]] = {}
    for r in refs:
        idx.setdefault(r["titulo"], []).append(r)
    return idx


def melhor_match_fuzzy(titulo: str, refs: list[dict], min_ratio: float = 0.87):
    """Para casos sem match exato, procura SequenceMatcher >= min_ratio."""
    melhor = None
    melhor_ratio = 0.0
    primeira_palavra = titulo.split()[0] if titulo else ""
    for r in refs:
        # Poda 1: comprimento muito diferente
        if abs(len(r["titulo"]) - len(titulo)) > len(titulo) * 0.3:
            continue
        # Poda 2: primeira palavra precisa bater (90% dos matches válidos)
        rt = r["titulo"]
        if rt.split()[0:1] != [primeira_palavra]:
            continue
        ratio = SequenceMatcher(None, titulo, rt).ratio()
        if ratio > melhor_ratio:
            melhor = r
            melhor_ratio = ratio
    if melhor_ratio >= min_ratio:
        return melhor, melhor_ratio
    return None, melhor_ratio


def cruzar(revisada: list[dict], refs: list[dict]) -> dict:
    idx = indexar_referencial(refs)
    log = Counter()
    resultados = []

    for rev in revisada:
        match = None
        confianca = None
        candidatos = idx.get(rev["titulo"], [])

        # Filtra candidatos com DOI canônico
        candidatos_com_doi = [c for c in candidatos if c["doi"]]

        if candidatos_com_doi:
            # Refina por ano + periódico
            por_ano_periodico = [
                c
                for c in candidatos_com_doi
                if c["ano"] == rev["ano"] and c["periodico"] == rev["periodico"]
            ]
            por_ano = [c for c in candidatos_com_doi if c["ano"] == rev["ano"]]

            if por_ano_periodico:
                match = por_ano_periodico[0]
                confianca = "alta_titulo_ano_periodico"
            elif por_ano:
                match = por_ano[0]
                confianca = "media_titulo_ano"
            else:
                match = candidatos_com_doi[0]
                confianca = "media_titulo_apenas"
        else:
            # Sem match exato: tenta fuzzy
            fuzzy, ratio = melhor_match_fuzzy(rev["titulo"], refs)
            if fuzzy and fuzzy["doi"]:
                # Reforça com ano se disponível
                if rev["ano"] and fuzzy["ano"] == rev["ano"]:
                    match = fuzzy
                    confianca = f"baixa_fuzzy_{ratio:.2f}_ano"
                else:
                    match = fuzzy
                    confianca = f"baixa_fuzzy_{ratio:.2f}"

        if match:
            log[f"match:{confianca}"] += 1
            resultados.append(
                {
                    "linha_xlsx_revisada": rev["linha_xlsx"],
                    "titulo_revisada": rev["titulo_orig"],
                    "ano_revisada": rev["ano"],
                    "periodico_revisada": rev["periodico"],
                    "doi_recuperado": match["doi"],
                    "confianca": confianca,
                    "titulo_referencial": match["titulo_orig"],
                    "linha_xlsx_referencial": match["linha_xlsx"],
                }
            )
        else:
            log["sem_match"] += 1

    return {"log": dict(log), "resultados": resultados}


def main(revisada_path: str, referencial_path: str, saida: str) -> None:
    revisada = carregar_revisada(Path(revisada_path))
    refs = carregar_referencial(Path(referencial_path))
    print(f"Revisada:    {len(revisada)} registros")
    print(f"Referencial: {len(refs)} registros (aba Base)")
    print(f"  com DOI canônico: {sum(1 for r in refs if r['doi'])}")

    cruzamento = cruzar(revisada, refs)

    payload = {
        "_meta": {
            "gerado_em": datetime.datetime.now(datetime.UTC).isoformat(timespec="seconds"),
            "revisada": Path(revisada_path).name,
            "referencial": Path(referencial_path).name,
            "total_revisada": len(revisada),
            "total_referencial": len(refs),
            "total_match": len(cruzamento["resultados"]),
            "log": cruzamento["log"],
            "criterios": {
                "alta_titulo_ano_periodico": "match exato em título normalizado + ano + periódico",
                "media_titulo_ano": "match exato em título + ano",
                "media_titulo_apenas": "match exato em título; ano/periódico divergem",
                "baixa_fuzzy_*": "SequenceMatcher ratio >= 0.92, sem match exato",
            },
        },
        "resultados": cruzamento["resultados"],
    }

    with open(saida, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"\nMatches encontrados: {len(cruzamento['resultados'])}/{len(revisada)}")
    for k, v in sorted(cruzamento["log"].items()):
        print(f"  {k:40s}  {v}")
    print(f"\nSalvo em: {saida}")


if __name__ == "__main__":
    if len(sys.argv) != 4:
        print(
            "Uso: recuperar_dois_referencial.py <revisada.xlsx> <referencial.xlsx> <saida.json>",
            file=sys.stderr,
        )
        sys.exit(2)
    main(sys.argv[1], sys.argv[2], sys.argv[3])
